from openpyxl import *
from tkinter import *
from tkinter import ttk, messagebox
import re

wb = load_workbook("D:\\Code\\Python\\2212453_NgoBaTai_Lab4\\dkhp.xlsx")
sheet = wb.active
years = ["2022-2023", "2023-2024", "2024-2025"]

def excel():
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 50
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 20
    sheet.column_dimensions["H"].width = 30

    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
    sheet.cell(row=1, column=88).value = "Môn học"


def focus1(event):
    name_field.focus_set()


def focus2(event):
    birthday_field.focus_set()


def focus3(event):
    email_id_field.focus_set()


def focus4(event):
    contact_no_field.focus_set()


def focus5(event):
    sem_field.focus_set()


def focus6(event):
    year_field.focus_set()


def focus7(event):
    year_field.focus_set()


def clear():
    id_field.delete(0, END)
    name_field.delete(0, END)
    sem_field.delete(0, END)
    birthday_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    year_field.delete(0, END)


def insert():
    id_ = id_field.get().strip()
    name = name_field.get().strip()
    sem = sem_field.get().strip()
    birthday = birthday_field.get().strip()
    contact_no = contact_no_field.get().strip()
    email = email_id_field.get().strip()
    year = year_field.get().strip()

    # Kiểm tra các trường có rỗng không
    if not (id_ and name and sem and birthday and contact_no and email and year):
        messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
        return

    # Kiểm tra định dạng dữ liệu
    if len(id_) < 7 or len(id_) > 8:
        messagebox.showerror("Lỗi", "Mã số sinh viên phải là số!")
        return
    if len(contact_no) < 10 or len(contact_no) > 11:
        messagebox.showerror("Lỗi", "Số điện thoại không hợp lệ!")
        return
    if not validate_email(email):
        messagebox.showerror("Lỗi", "Email không hợp lệ!")
        return
    if not is_valid_date(birthday):
        messagebox.showerror("Lỗi", "Ngày sinh phải có dạng dd/mm/yyyy!")
        return

    # Kiểm tra có chọn môn học chưa
    selected_subjects = [subj for subj, var in subjects.items() if var.get()]
    if not selected_subjects:
        messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất một môn học!")
        return

    # Ghi vào Excel
    try:
        file_path = "D:\\Code\\Python\\2212453_NgoBaTai_Lab4\\dkhp.xlsx"
        wb = load_workbook(file_path)
        sheet = wb.active

        for subject in selected_subjects:
            row = sheet.max_row + 1
            sheet.append([id_, name, birthday, email, contact_no, sem, year, subject])

        wb.save(file_path)
        messagebox.showinfo("Thành công", "Đăng ký thành công!")

        # Xóa nội dung sau khi lưu
        clear()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể ghi file: {e}")


def is_valid_date(date):
    pattern = r"^\d{2}/\d{2}/\d{4}$"  # Định dạng dd/mm/yyyy
    return re.fullmatch(pattern, date) is not None


def validate_input(new_value):
    if new_value.isdigit() or new_value == "":
        return True
    return False   

def validate_semester_input(new_value):
    if new_value.isdigit() and int(new_value) <= 3 or new_value == "":
        return True
    return False


def validate_email(new_value):
    pattern = r"^[a-zA-Z0-9_.@+-]*$"  # Cho phép các ký tự trong email
    return re.fullmatch(pattern, new_value) is not None


if __name__ == "__main__":
    root = Tk()
    root.configure(background="light green")
    root.title("Form đăng kí học phần")
    root.geometry("500x400")
    excel()

    heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green",fg = "red", font=("Arial", 18, "bold"))
    id = Label(root, text= "Mã số sinh viên", bg="light green")
    name = Label(root, text="Họ tên", bg="light green")
    birthday = Label(root, text="Ngày sinh", bg="light green")
    sem = Label(root, text="Học kỳ", bg="light green")
    year = Label(root, text="Năm học", bg="light green")
    contact_no = Label(root, text="Số điện thoại", bg="light green")
    email_id = Label(root, text="Email", bg="light green")
    study_id = Label(root, text="Chọn môn học", bg = "light green")

    heading.grid(row=0, column=1)
    id.grid(row=1, column=0)
    name.grid(row=2, column=0)
    birthday.grid(row=3, column=0)
    email_id.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    sem.grid(row=6, column=0)
    year.grid(row=7, column=0)
    study_id.grid(row=8, column = 0)

    # ---Kiểm tra chỉ cho nhập số
    vcmd = root.register(validate_input)
    email = root.register(validate_email)
    semester = root.register(validate_semester_input)
    # ---------------------------
    id_field = Entry(root, validate="key", validatecommand=(vcmd, "%P"))
    name_field = Entry(root)
    sem_field = Entry(root, validate="key", validatecommand=(semester, "%P"))
    birthday_field = Entry(root)
    contact_no_field = Entry(root, validate="key", validatecommand=(vcmd, "%P"))
    email_id_field = Entry(root, validate="key", validatecommand=(email, "%P"))
    year_field = ttk.Combobox(root, values=years, state="readonly")

    id_field.bind("<Return>", focus1)
    name_field.bind("<Return>", focus2)
    birthday_field.bind("<Return>", focus3)
    email_id_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    sem_field.bind("<Return>", focus6)
    year_field.bind("<Return>", focus7)

    id_field.grid(row = 1, column=1, ipadx="100")
    name_field.grid(row=2, column=1, ipadx="100")
    birthday_field.grid(row=3, column=1, ipadx="100")
    email_id_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    sem_field.grid(row=6, column=1, ipadx="100")
    year_field.grid(row=7, column=1, ipadx="92")
    

    frame_checkbox = Frame(root, bg="light green")
    frame_checkbox.grid(row=9, column=1)

    subjects = {
        "Python": IntVar(),
        "Java": IntVar(),
        "C": IntVar(),
        "Web": IntVar(),
    }
    python = IntVar()
    checkbox1 = Checkbutton(frame_checkbox, text="Lập trình python", variable=subjects["Python"], bg="light green")
    java = IntVar()
    checkbox2 = Checkbutton(
        frame_checkbox, text="Lập trình Java", variable=subjects["Java"], bg="light green"
    )
    tech = IntVar()
    checkbox3 = Checkbutton(
        frame_checkbox,
        text="Công nghệ phần mềm",
        variable=subjects["C"],
        bg="light green",
    )
    web = IntVar()
    checkbox4 = Checkbutton(
        frame_checkbox,
        text="Phát triển ứng dụng web",
        variable=subjects["Web"],
        bg="light green",
    )

    checkbox1.grid(row=0, column=0, padx=5, sticky="w")
    checkbox2.grid(row=0, column=1, padx=5, sticky="w")
    checkbox3.grid(row=1, column=0, padx=5, sticky="w")
    checkbox4.grid(row=1, column=1, padx=5, sticky="w")

    frame_buttons = Frame(root, bg="light green", pady=10)
    frame_buttons.grid(row=10, column=1)  

    excel()
    submit = Button(frame_buttons, text="Đăng ký", fg="Black", bg="Green", command=insert)
    submit.pack(side = LEFT, padx=75)   

    cancel = Button(frame_buttons, text="Thoát", fg="Black", bg="Green", command=quit)
    cancel.pack(side = RIGHT, padx=50)

    root.mainloop()
